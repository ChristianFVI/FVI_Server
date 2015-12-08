__author__ = 'Joshua Menke'
import json
import os
import fnmatch
import time
from threading import Thread
from tinify import tinify
import re
from xml.dom import minidom
import shutil

from openpyxl import load_workbook
from PIL import Image
import bottlenose
import requests
from bs4 import BeautifulSoup
from openpyxl.comments import Comment

print(time.strftime('%H:%M - %d.%m.%y'))
crnt = os.path.dirname(os.path.realpath(__file__))
print(crnt)
#CHANGES IN UNIX MERGE
def copyDirectory(src, dest):
    try:
        shutil.copytree(src, dest)
    # Directories are the same
    except shutil.Error as e:
        print('Directory not copied. Error: %s' % e)
    # Any error saying that the directory doesn't exist
    except OSError as e:
        print('Directory not copied. Error: %s' % e)

home = os.path.expanduser("~")
with open(crnt + r"/gg.json", "r") as file:
    gg = json.load(file)
tinify.key = "BmMHiAQQI8mOJhhMasMuojog_B0LJqGV"
with open(crnt + r"/gg.json", "w") as file:
    json.dump(gg, file)
bottlenose.Amazon()
amazon = bottlenose.Amazon("AKIAIL6WN32WBBYNIY4Q", "PchxA1xqaaqRseDKJWxW0ZKhV8elJIBcMaUWqJHJ", "topse07-21", MaxQPS=1.2,
                           Region="DE")

def hasNumbers(inputString):
    return any(char.isdigit() for char in inputString)
# amazon = bottlenose.Amazon(MaxQPS=0.8)
def move_over(src_dir, dest_dir):
    fileList = os.listdir(src_dir)
    for i in fileList:
        src = os.path.join(src_dir, i)
        dest = os.path.join(dest_dir, i)
        if os.path.exists(dest):
            if os.path.isdir(dest):
                move_over(src, dest)
                continue
            else:
                os.remove(dest)
        shutil.move(src, dest_dir)


def resize3(path):
    img = Image.open(path)
    img_w, img_h = img.size
    bg_s = img_w
    if img_h > bg_s:
        bg_s = img_h
    background = Image.new("RGBA", (bg_s, bg_s), (255, 255, 255, 255))
    offset = (int((bg_s - img_w) / 2), int((bg_s - img_h) / 2))
    background.paste(img, offset)
    background = background.resize((1000, 1000))
    background.save(path)
def resize2(path):
    img = Image.open(path)
    maxsize = (1000, 1000)
    img.thumbnail(maxsize)
    img.save(path)

def resize(path):
    img = Image.open(path)
    img_w, img_h = img.size
    bg_s = img_w
    if img_h > bg_s:
        bg_s = img_h
    background = Image.new("RGBA", (bg_s, bg_s), (255, 255, 255, 255))
    offset = (int((bg_s - img_w) / 2), int((bg_s - img_h) / 2))
    background.paste(img, offset)
    if bg_s > 1000:
        background = background.resize((1000, 1000))
    background.save(path)


def get_edit(filename):
    with open(filename, "r") as file:
        return json.load(file)


def sav_edit(filename, edit):
    with open(filename, "w")as jfile:
        json.dump(edit, jfile)


class WatchResize(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.deamon = True
        self.start()

    def run(self):
        date = time.strftime('%H:%M - %d.%m.%y')
        path_to_watch = home + "/Google Drive/TinyPng/Resize It/"
        path_to_write = home + "/Google Drive/TinyPng/Resized/"
        before = dict([(f, None) for f in os.listdir(path_to_watch)])
        while 1:
            edited = get_edit("Resi.json")
            after = dict([(f, None) for f in os.listdir(path_to_watch)])
            added = [f for f in after if not f in before]
            removed = [f for f in before if not f in after]
            # if added:
            for i in after:
                ending = os.path.splitext(i)[1]
                if ending == ".jpg" or ending == ".png" or ending == ".jpeg" or ending == ".gif" or ending == ".PNG":
                    path = path_to_watch + i
                    gname = os.path.abspath(path)
                    v = os.path.splitext(i)[0]
                    if gname not in edited:
                        try:
                            source = tinify.from_file(path)
                            source.to_file(path_to_write + v + ".png")
                        except:
                            tinify.key = gg.remove(tinify.key)
                            tinify.key = gg[0]
                            with open(crnt + r"/gg.json", "w") as file:
                                json.dump(gg, file)

                        resize(path_to_write + v + ".png")
                        edited.append(gname)
                    os.remove(path)
            if added:
                print("Added: ", ", ".join(added))

            if removed: print("Removed: ", ", ".join(removed))
            before = after
            sav_edit("Resi.json", edited)


class WatchPic(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.deamon = True
        self.start()

    def run(self):
        def prodbilder(prodkat, pn):
            er = False
            try:

                for filename in find_files(home + r"/server@fvi.rocks/Produktbilder/nicht editiert/" + prodkat,
                                           '*.' + pn):
                    pname = filename.replace(home + r"/server@fvi.rocks/Produktbilder/nicht editiert/" + prodkat, "")
                    gname = os.path.abspath(filename)
                    try:
                        resize(gname)
                        try:
                            source = tinify.from_file(gname)
                            source.to_file(gname)
                        except Exception as e:
                            f = open(home + r"/server@fvi.rocks/Produktbilder/fehler/error.log", 'a')
                            for g in e.args:
                                print("\a")
                                f.write(g)
                                f.write(pname + ": " + "\n")
                            f.close()
                        print(pname)

                    except Exception as e:
                        er = True
                        print("Error: " + pname)
                        print(e)
            except Exception as e:
                er = True
                print(e)
            return er

        while 1:
            path_to_watch = home + "/server@fvi.rocks/Produktbilder/nicht editiert"
            after = os.listdir(path_to_watch)
            for i in after:
                if os.path.isdir(path_to_watch + "/" + i) and os.listdir(path_to_watch + "/" + i):
                    try:
                        e = []
                        e.append(prodbilder(i, "jpg"))
                        e.append(prodbilder(i, "png"))
                        e.append(prodbilder(i, "jpeg"))
                        e.append(prodbilder(i, "jfif"))
                        e.append(prodbilder(i, "jpe"))
                        src = home + r"/server@fvi.rocks/Produktbilder/nicht editiert/" + i
                        date = time.strftime('%H.%M - %d.%m.%y')
                        dst = home + r"/server@fvi.rocks/Produktbilder/fertig editiert/" + i + "(" + date + ")"
                        if True in e:
                            dst = home + r"/server@fvi.rocks/Produktbilder/fehler/" + i + "(" + date + ")"
                        shutil.move(src, dst)
                    except Exception as e:
                        print(e)
            time.sleep(10)


class WatchShop(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.deamon = True
        self.start()

    def run(self):
        def prodbilder(prodkat, pn):
            er = False
            try:

                for filename in find_files(
                                        home + r"/server@fvi.rocks/Produktbilder/Shopbilder nicht editiert/" + prodkat,
                                        '*.' + pn):
                    pname = filename.replace(
                        home + r"/server@fvi.rocks/Produktbilder/Shopbilder nicht editiert/" + prodkat, "")
                    gname = os.path.abspath(filename)
                    try:
                        resize2(gname)
                        try:
                            source = tinify.from_file(gname)
                            source.to_file(gname)
                        except Exception as e:
                            f = open(home + r"/server@fvi.rocks/Produktbilder/fehler/error.log", 'a')
                            for g in e.args:
                                print("\a")
                                f.write(g)
                                f.write(pname + ": " + "\n")
                            f.close()
                        print(pname)
                    except Exception as e:
                        er = True
                        print("Error: " + pname)
                        print(e)
            except Exception as e:
                er = True
                print(e)
            return er

        while 1:
            path_to_watch = home + "/server@fvi.rocks/Produktbilder/Shopbilder nicht editiert"
            after = os.listdir(path_to_watch)
            for i in after:
                if os.path.isdir(path_to_watch + "/" + i) and os.listdir(path_to_watch + "/" + i):
                    try:
                        e = []
                        e.append(prodbilder(i, "jpg"))
                        e.append(prodbilder(i, "png"))
                        e.append(prodbilder(i, "jpeg"))
                        e.append(prodbilder(i, "jfif"))
                        e.append(prodbilder(i, "jpe"))
                        date = time.strftime('%H.%M - %d.%m.%y')
                        src = home + r"/server@fvi.rocks/Produktbilder/Shopbilder nicht editiert/" + i
                        dst = home + r"/server@fvi.rocks/Produktbilder/Shopbilder fertig editiert/" + i + "(" + date + ")"
                        if True in e:
                            dst = home + r"/server@fvi.rocks/Produktbilder/fehler/" + i + "(" + date + ")"

                        shutil.move(src, dst)
                    except Exception as e:
                        print(e)
            time.sleep(10)


class WatchScreen(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.deamon = True
        self.start()

    def run(self):
        def prodbilder(prodkat, pn):
            er = False
            try:

                for filename in find_files(
                                        home + r"/server@fvi.rocks/Produktbilder/Shopscreenshots nicht editiert/" + prodkat,
                                        '*.' + pn):
                    pname = filename.replace(
                        home + r"/server@fvi.rocks/Produktbilder/Shopscreenshots nicht editiert/" + prodkat, "")
                    gname = os.path.abspath(filename)
                    try:
                        resize3(gname)
                        try:
                            source = tinify.from_file(gname)
                            source.to_file(gname)
                        except Exception as e:
                            f = open(home + r"/server@fvi.rocks/Produktbilder/fehler/error.log", 'a')
                            for g in e.args:
                                print("\a")
                                f.write(g)
                                f.write(pname + ": " + "\n")
                            f.close()
                        print(pname)
                    except Exception as e:
                        er = True
                        print("Error: " + pname)
                        print(e)
            except Exception as e:
                er = True
                print(e)
            return er

        while 1:
            path_to_watch = home + "/server@fvi.rocks/Produktbilder/Shopscreenshots nicht editiert"
            after = os.listdir(path_to_watch)
            for i in after:
                if os.path.isdir(path_to_watch + "/" + i) and os.listdir(path_to_watch + "/" + i):
                    try:
                        e = []
                        e.append(prodbilder(i, "jpg"))
                        e.append(prodbilder(i, "png"))
                        e.append(prodbilder(i, "jpeg"))
                        e.append(prodbilder(i, "jfif"))
                        e.append(prodbilder(i, "jpe"))
                        date = time.strftime('%H.%M - %d.%m.%y')
                        src = home + r"/server@fvi.rocks/Produktbilder/Shopscreenshots nicht editiert/" + i
                        dst = home + r"/server@fvi.rocks/Produktbilder/Shopscreenshots fertig editiert/" + i + "(" + date + ")"
                        if True in e:
                            dst = home + r"/server@fvi.rocks/Produktbilder/fehler/" + i + "(" + date + ")"

                        shutil.move(src, dst)
                    except Exception as e:
                        print(e)
            time.sleep(10)


# t

class WatchProd(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.deamon = True
        self.start()

    def run(self):
        while 1:
            path_to_watch = home + "/server@fvi.rocks/Tabellen/Zu berechnende Tabellen"
            after = os.listdir(path_to_watch)
            for i in after:
                print(i)
                if not os.path.isdir(path_to_watch + "/" + i):
                    try:
                        ubergabe(home + r"/server@fvi.rocks/Tabellen/Zu berechnende Tabellen/" + i)
                    except Exception as e:
                        print("Error:")
                        print(e)
                        f = open(home + r"/server@fvi.rocks/Tabellen/Fehlerhafte Tabellen/error.log", 'a')
                        for g in e.args:
                            print("\a")
                            f.write(g)
                            f.write("\n")
                        f.close()
                        time.sleep(3)
                        try:

                            if os.path.exists(home + r"/server@fvi.rocks/Tabellen/Fehlerhafte Tabellen/" + i):
                                os.remove(home + r"/server@fvi.rocks/Tabellen/Fehlerhafte Tabellen/" + i)
                            shutil.move(home + r"/server@fvi.rocks/Tabellen/Zu berechnende Tabellen/" + i,
                                        home + r"/server@fvi.rocks/Tabellen/Fehlerhafte Tabellen/")
                            if os.path.exists(
                                                    home + r"/server@fvi.rocks/Tabellen/Zu berechnende Tabellen/" + i) and os.path.exists(
                                                    home + r"/server@fvi.rocks/Tabellen/Fehlerhafte Tabellen/" + i):
                                os.remove(home + r"/server@fvi.rocks/Tabellen/Zu berechnende Tabellen/" + i)
                        except Exception as e:
                            print(e)
                            f = open(home + r"/server@fvi.rocks/Tabellen/Fehlerhafte Tabellen/error.log", 'a')
                            for g in e.args:
                                print("\a")
                                f.write(i + ": " + g)
                                f.write("\n")
                            f.close()
            time.sleep(10)


def bstbe(i):
    if i <= 26:
        return chr(ord("A") + i - 1)
    else:
        p = i - 26
        return "A" + str(chr(ord("A") + p - 1))
def find_files(directory, pattern):
    for root, dirs, files in os.walk(directory):
        for basename in files:
            if fnmatch.fnmatch(basename, pattern):
                filename = os.path.join(root, basename)
                yield filename


def newtable():
    '''for filename in find_files(home + r"\Google Drive\Tabellen\Komplett Fertig", '*.xlsx'):
        try:
            gname = os.path.basename(filename)
            print(gname)
            sample = load_workbook("Musterblatt.xlsx")
            ss = sample.active
            wb = load_workbook(filename)
            ws = wb.active
            #Wertungen:
            ss["B2"] = ws["C2"].value
            ss["B3"] = ws["F2"].value
            ss["B4"] = ws["I2"].value
            #Tabvars
            ss["B7"] = ws["M2"].value
            ss["B8"] = ws["Q2"].value
            ss["B9"] = ws["U2"].value
            ss["B10"] = ws["Y2"].value
            ss["B11"] = ws["AC2"].value
            ss["B12"] = ws["AG2"].value
            #Produktnamen
            ss2 = sample.get_sheet_by_name("Produktdaten")
            for i in range(10):
                print(i)
                p = chr(ord("B") + i)
                ss2[p+"1"] = ws["A"+str(i+2)].value
            #Tabwert1
            for p in range(6):
                bstbe = chr(ord("N") + 4 * int(p))
                if p == 4:
                    bstbe = "AD"
                if p == 5:
                    bstbe = "AH"
                print(bstbe)
                for i in range(10):
                    d = chr(ord("B") + int(i))
                    ss2[d+str(p+4)] = ws[bstbe+str(i+2)].value
            #ASIN
            for i in range(10):
                p = chr(ord("B") + i)
                ss2[p+"2"] = ws["AX"+str(i+2)].value

            sample.save(home + r"\Google Drive\Tabellen (2)\Zu berechnende Tabellen\\"+gname)
        except:
    '''
def ubergabe(path):

    ping = path
    # print("We got here")

    wb = load_workbook(path)
    mb = load_workbook(crnt + r"/Abzulesen (DATUM).xlsx")
    ws1 = wb.get_sheet_by_name("Variablen")
    ws2 = wb.get_sheet_by_name("Produktdaten")
    ms1 = mb.active
    prod = True
    nr = 0
    Wertungen = []
    '''def (i):
        if "%" in i:
            p  = re.sub("[\s%]+","",i)   Przntberechnung
            return int(p) / 100
    #print(("10%"))'''

    Wertungen.append(ws1["B2"].value)
    Wertungen.append(ws1["B3"].value)
    Wertungen.append(ws1["B4"].value)
    ms1["A5"] = ws1["B2"].value
    ms1["A6"] = ws1["B3"].value
    ms1["A7"] = ws1["B4"].value
    asin = []

    # Tabvars
    tabvar = []
    tab = True
    while True:
        if ws1["B" + (str(7 + nr))].value != None:
            ms1["A" + str(8 + nr)] = ws1["B" + (str(7 + nr))].value
            tabvar.append(ws1["B" + (str(7 + nr))].value)
        else:
            break
        nr += 1
    nr2 = 0
    while True:
        b = bstbe(nr2+2)
        if ws2[b + "1"].value != None:
            ms1[b + "1"] = ws2[b + "1"].value
        else:
            break
        nr2 += 1
    # WERTUNG1
    nr3 = 0
    wert1 = []
    while True:
        print(nr3)
        b = bstbe( 4+ nr3 * 3)
        wert1.append(ws1[b + "2"].value)
        if wert1[nr3] == None:
            wert1.remove(None)
            break
        nr3 += 1
    gewicht1 = 1 / nr3
    #print(wert1)
    # WERTUNG2
    nr4 = 0
    wert2 = []
    while True:
        b = bstbe(4 + nr4 * 3)
        wert2.append(ws1[b + "3"].value)
        if wert2[nr4] == None:
            wert2.remove(None)
            break
        nr4 += 1
    gewicht2 = 1 / nr4
    #print(wert2)
    # WERTUNG3
    nr5 = 0
    wert3 = []
    while True:
        if nr5 <=7:
            b = chr(ord("D") + nr5 * 3)
        else:
            b = "A" + str(chr(ord("B")+(nr5-8)*3))
        print(b)
        wert3.append(ws1[b + str(4)].value)
        if wert3[nr5] == None:
            wert3.remove(None)
            break

        nr5 += 1
    gewicht3 = 1 / nr5
    #print(wert3)
    # GEWICHTUNG1
    gew1 = []
    for i in range(len(wert1)):

        b = bstbe(6 + i * 3)
        print(ws1[b + "2"].value)
        gew1.append((ws1[b + "2"].value))
        print("BIS HIER HIN!")
        if gew1[0] == None:
            gew1.remove(None)
    # if gew1.index(None)!=-1:
    #    gew1.remove(None)
    if len(gew1) == 0:
        for i in range(len(wert1)):
            gew1.append(gewicht1)
    #print(gew1)
    # GEWICHTUNG2
    gew2 = []
    for i in range(len(wert2)):
        # print(i)
        b = bstbe(6 + i * 3)
        gew2.append((ws1[b + "3"].value))
        if gew2[0] == None:
            gew2.remove(None)
            # if gew2.index(None)!=-1:
            #   gew2.remove(None)
    if len(gew2) == 0:
        for i in range(len(wert2)):
            gew2.append(gewicht2)
    #print(gew2)
    # GEWICHTUNG3
    gew3 = []
    for i in range(len(wert3)):

        b = bstbe(6 + i * 3)
        gew3.append((ws1[b + "4"].value))
        if gew3[0] == None:
            gew3.remove(None)
    # if gew3.index(None)!=-1:
    #    gew3.remove(None)
    if len(gew3) == 0:
        for i in range(len(wert3)):
            gew3.append(gewicht3)
    #print(gew3)
    # Plus Minus
    plu1 = []
    for i in range(len(wert1)):
        # print(i)
        b = bstbe(5 + i * 3)
        if ws1[b + "2"].value != None:
            plu1.append(ws1[b + "2"].value)
        else:
            plu1.append("+")
    #print(plu1)
    # Plus Minus 2
    plu2 = []
    for i in range(len(wert2)):
        # print(i)
        b = bstbe(5 + i * 3)
        if ws1[b + "3"].value != None:
            plu2.append(ws1[b + "3"].value)
        else:
            plu2.append("+")
    #print(plu2)
    # Plus Minus 3
    plu3 = []
    for i in range(len(wert3)):
        # print(i)
        b = bstbe(5 + i * 3)
        if ws1[b + "4"].value != None:
            plu3.append(ws1[b + "4"].value)
        else:
            plu3.append("+")
    #print(plu3)
    ##############################
    #print("stop")
    for i in range(nr2):  # Prudktnamen
        #print(i)
        #print("SS")
        b = bstbe(2 + i)
        ms1[b + "4"] = ws2[b + "3"].value
        for j in range(nr):  # Tabvars
            ms1[b + str(j + 8)] = ws2[b + str(j + 4)].value

    def get_amazon_star(asin,c):

        try:
            r = requests.get("http://www.amazon.de/gp/product/"+asin)
            time.sleep(1)
            asd = r.text
            s = BeautifulSoup(r.text,"html.parser")
            # time.sleep(1)
            g = s.find('div', id='avgRating')
            m = re.search(r'\d+(\.\d+)?', g.text.strip())
            f = float(m.group(0))
            g = f / 5 * 100
            return g
        except:
            #time.sleep(1)
            c = c+1
            if c == 3:
                return 4 / 5 *100
            print("Waiting for Amazon")
            return get_amazon_star(asin,c)
    def getabs(i, j):
        merk = []
        cnt = 0
        j = int(j) + 3
        while True:
            b = bstbe(3 + cnt)

            if ws1[b + str(j)].value != None:
                merk.append(ws1[b + str(j)].value)
            else:
                break
            cnt += 1
        print("ABSOLUT VON:")
        print(i)
        print("Reihe:")
        print(j)
        print(merk)
        if len(merk) != 0:
            gewichtung = 1 / len(merk)
            cnt = 0
            for g in merk:
                if i == g:
                    return 100 - 100 * gewichtung * cnt
                else:
                    cnt += 1
        if i == "Ja" or i == "ja":
            return 100
        elif i == "Nein" or i == "nein":
            return 50
        ret = str(i)
        ret = re.sub('\.', "", ret)
        ret = re.sub('X+', "*", ret)
        ret = re.sub('x+', "*", ret)
        ret = re.sub('\,+', '.', ret)
        ret = re.sub('(?![\*\.\-\/+])\D', "", ret)
        ret = re.sub('[\/\*\+\-](?!\d)', "", ret)
        # print(ret)
        if hasNumbers(ret):
            try:
                ret = eval(ret)
            except:
                ret = 1
        else:
            ret = 0
        print("ABSOLUT: ################")
        print(ret)
        # print("JOP")
        return ret
    amazon_p = []
    if "Amazon" in wert1 or "Amazon" in wert2 or "Amazon" in wert3:
        for i in range(nr2):
            b = bstbe(2 + i)
            a = ws2[b+str(2)].value
            asin.append(a)
            g = get_amazon_star(a,0)
            amazon_p.append(g)
            #time.sleep(1)
            #print(r.text)
            #xmldoc = minidom.parseString(r.text)
            #itemlist = xmldoc.getElementsByTagName("div")
            #print(itemlist)
    tabwerte = {}
    tabwerte["Amazon_p"] = amazon_p
    for i in range(len(tabvar)):
        print(i)
        e = i + 4
        d = tabvar[i]
        #p = re.sub("\s","",i)
        tabwerte[d] = []
        for j in range(nr2):
            b = bstbe(2 + j)
            tabwerte[d].append(getabs(ws2[b+str(e)].value,e))
        tabwerte[d+"_s"] = sorted(tabwerte[d])
        l =round(len(tabwerte[d+"_s"])/2)
        median = (tabwerte[d+"_s"][l-1] + tabwerte[d+"_s"][l])/2
        max = median*1.5
        if max == 0:
            max = 0.1
        tabwerte[d+"_p"] = []
        print("Zeile 746")
        for f in tabwerte[d]:
            print(f)
            print(max)
            if f != 0:

                value = f / max * 100
            else:
                value = 0
            if value > 100:
                value = 100
            if value < 0:
                value = 0
            tabwerte[d+"_p"].append(value)
    print("Zeile 758")
    print (tabvar)
    print (tabwerte)
    def get_star(g, wert, v):
        tabwert = []
        for i in range(nr2):
            b = bstbe(2 + i)
            tabwert.append(getabs(ws2[b + str(g + 4)].value, v))
        tabwert.sort()
        l = round((len(tabwert)) / 2)
        median = (tabwert[l - 1] + tabwert[l]) / 2
        max = tabwert[len(tabwert)-1]
        #wert = getabs(wert, v)
        print (str(max) + "    " + str(wert))
        value = wert / max * 100
        if value > 100:
            value = 100
        if value < 0:
            value = 0
        print (tabwert)
        print(v)
        print(wert)
        return value

    def stern(wert, gewt, plus, bst):
        add = []
        for i in wert:

            add.append(i)
        #print (add)
        sume = 0
        a = 0
        gesgew = 0
        for i in add:
            v = i
            print("JETZ KOMMT A:")
            print(a)
            print(bst)
            s = ord(bst) - ord("B")
            print(s)
            d = tabwerte[i+"_p"][s]
            if plus[a] == "-":
                d = 100 - d
            gesgew = gesgew + gewt[a]
            sume = sume + gewt[a] * d
            a += 1
        print("Und die summe:")
        print(sume)
        return sume / gesgew

    # print(getabs("Gummi",7))
    # print("TT")
    # print(stern(wert2,gew2,plu2,"C"))
    def amazon_price(asin):
        xml = amazon.ItemLookup(ItemId=asin, ResponseGroup="Offers")
        xmldoc = minidom.parseString(xml)
        itemlist = xmldoc.getElementsByTagName("Price")
        print (xml)
        try:
            price = itemlist[0].lastChild.firstChild.nodeValue
        except:
            price = "0"
        # price = itemlist.getElementsByTagName("Formatted Price")
        # itemlist2 = xmldoc.toprettyxml()
        return price

    #print(amazon_price("B003TSFDM0"))

    def endnote(w1,w2,w3,h):
        #stern1 = stern(wert1, gew1, plu1, h)
        stern1 = w1
        a = (ws1["C2"].value)
        if a == None:
            a = 1 / 3

        #print("#######")
        #print(stern1)
        #print(wert1)
        #print(gew1)
        #print(plu1)
        #print(b)
        #print("#######")
        #stern2 = stern(wert2, gew2, plu2, h)
        stern2 = w2
        b = (ws1["C3"].value)
        if b == None:
            b = 1 / 3

        #print("#######")
        #print(stern2)
        #print(wert2)
        #print(gew2)
        #print(plu2)
        #print(b)
        #print("#######")
        #stern3 = stern(wert3, gew3, plu3, h)
        stern3 = w3
        c = (ws1["C4"].value)
        if c == None:
            c = 1 / 3
        #print("#######")
        #print(stern3)
        #print(wert3)
        #print(gew3)
        #print(plu3)
        #print(b)
        #print("#######")
        print(a)
        print(b)
        print(c)
        endstern = (stern1 * a + stern2 * b + stern3 * c) / (a + b + c)
        print("Hier kommt der Endstern")
        print(endstern)
        for i in range(100,0,-3):
            t = 100 + i * -4
            if endstern >= i:
                #print(i)
                p = 1.0 + 0.1 * i
                return (0.9 + ((100 - i) /3 * 0.1))
                #return p
        return 2.5
    #Sterne
    wert1n = []
    wert2n = []
    wert3n = []
    endnoten = []
    for i in range(nr2):
        b = bstbe(2 + i)
        aa = stern(wert1,gew1,plu1,b)
        bb = stern(wert2,gew2,plu2,b)
        cc = stern(wert3,gew3,plu3,b)
        wert1n.append(aa)
        wert2n.append(bb)
        wert3n.append(cc)
        endnoten.append(endnote(aa,bb,cc,b))
    print(sum(endnoten))
    print("HIER KOMMEN ENDNOTEN:")
    print(endnoten)
    sortend = sorted(endnoten)
    p = 2
    if ws2["B2"].value == None:
        p = 1.8
    if sortend[0] >= p:
        mp = (p-0.1) / sortend[0]
        for i in range(len(endnoten)):
            endnoten[i] = round(endnoten[i] * mp,1)
    sortend = sorted(endnoten)
    print(endnoten)
    print (wert1n)
    def get_stern(b):
        g = 18
        if ws2["B2"].value == None or ws2["B2"].value == "":
            g = 22
        for i in range(100,0,-g):
            print(i)
            t = 100 + i * -100/6
            if b > i:
                d = 5 - i *0.5
                d = (5.5-((100-i)/g*0.5))
                return d
        return 2.5
    preisleist = []
    p = len(tabvar) + 8
    ms1["A"+str(p)]="ASIN"
    for i in range(nr2):
        b = bstbe(2 + i)
        ms1[b+str(2)]=endnoten[i]
        print("HIER KOMMT STERN 1 von Reihe"+b)
        print(wert1n[i])
        ms1[b+str(5)]=get_stern(wert1n[i])
        print("HIER KOMMT STERN 2 von Reihe"+b)
        print(wert2n[i])
        ms1[b+str(6)]=get_stern(wert2n[i])
        print("HIER KOMMT STERN 3 von Reihe"+b)
        print(wert3n[i])
        ms1[b+str(7)]=get_stern(wert3n[i])
        if ws2["B2"].value != None:
            preis = amazon_price(ws2[b+str(2)].value)
            ms1[b+str(p)]=ws2[b+str(2)].value
            ret = re.sub('\.', "", preis)
            ret = re.sub('X+', "*", ret)
            ret = re.sub('x+', "*", ret)
            ret = re.sub('\,+', '.', ret)
            ret = re.sub('(?![\*\.\-\/+])\D', "", ret)
            ret = re.sub('[\/\*\+\-](?!\d)', "", ret)
            print(ret)
            preisleist.append(float(ret)*endnoten[i])

    for i in range(len(tabvar)):
        b = "M"
        t = 8 + i
        f = 7 + i
        if ws1[b+str(f)].value != None:

            comment = Comment(str(ws1[b+str(f)].value),"Joshua")
            ms1["A"+str(t)].comment = comment
    preisleister = 1000000
    testsieger = 10000
    for i in range(len(endnoten)):
        if endnoten[i] != 0 and endnoten[i] < testsieger:
            testsieger = endnoten[i]

    for i in range(nr2):
        b = bstbe(2 + i)
        if endnoten[i] == testsieger:
            ms1[b+str(3)]="Testsieger"
            break
    drop = []
    if ws2["B2"].value != None:
        for j in range(4):
            drop.append(sortend[j])
    if ws2["B2"].value != None:
        for i in range(len(endnoten)):
            b = bstbe(2 + i)
            # 1
            print("ENDNOOOOTTEEENNENENNENENENENNENENE:")
            print(drop)
            if endnoten[i] in drop:
                if preisleist[i] != 0 and preisleist[i] < preisleister and ms1[b+str(3)].value != "Testsieger":
                    preisleister = preisleist[i]
        for i in range(nr2):
            b = bstbe(2 + i)
            if preisleist[i] == preisleister:
                ms1[b+str(3)]="Preisleistungssieger"
        print(preisleister)
        print(testsieger)
        print(preisleist)
    savepath = os.path.basename(ping)
    savepath = os.path.splitext(savepath)[0]
    print(savepath)
    print (wert1n)
    print(wert2n)
    print(wert3n)
    date = time.strftime('%H.%M - %d.%m.%y')
    print(date)
    print(wert1)
    print(wert1n)

    print(gew1)
    mb.save(home + r"/server@fvi.rocks/Tabellen/Auszulesende Tabellen/" + savepath + " (" + date + ").xlsx")
    path2 = os.path.splitext(path)[0] + " (" + date + ").xlsx"
    print(path)
    wb.save(path2)
    time.sleep(1)
    if os.path.exists(home + r"/server@fvi.rocks/Tabellen/Berechnete Tabellen/" + savepath + " (" + date + ").xlsx"):
        os.remove(home + "/server@fvi.rocks/Tabellen/Berechnete Tabellen/" + savepath + " (" + date + ").xlsx")
    shutil.move(path2, home + r"/server@fvi.rocks/Tabellen/Berechnete Tabellen/")
    os.remove(path)
    #print(endnote("B"))


WatchScreen()
WatchPic()
WatchShop()
WatchProd()

while True:
    pass
